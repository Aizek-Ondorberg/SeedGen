"""
Программа генерация сида робота:
    - считывает ФИО из файла people.xlsx
    - генерирует случайный не повторяющийся сид робота
    - записывеет ФИО и сиды в файл spisok.xlsx
Авторы:
Студенты группы КРБО-О3-17
Горовиц К.А.
Тихонов Н.С.
Тихонов С.И.

Создано в PyCharm Professional 2020.3
Build #PY-203.7148.72, build on January 27, 2021

"""
import random
import openpyxl
import xlsxwriter as xw
from collections import Counter
g=0
mint= ''
shnaps = []
sovpad = []
surname = []
T = 0
for o in range(1, 30):
    wrb = openpyxl.reader.excel.load_workbook(filename="people.xlsx", data_only=True)
    wrb.active = 0
    sheet_f = wrb.active
    se = sheet_f['A' + str(o)].value
    surname.append(se)
workbook = xw.Workbook('spisok.xlsx')
worksheet = workbook.add_worksheet("Лист1")
im = ''
for t in range(0, 29):
    worksheet.write(t+1, 0, surname[t - 1])


    while len( sovpad ) != len(surname) : # пока длина массива с сидами не равна длине массива с фио выполянем генерацию сида
            wb = openpyxl.reader.excel.load_workbook(filename="Robot.xlsx", data_only=True)
            wb_f = openpyxl.reader.excel.load_workbook(filename="Robot.xlsx")
            wb.active = 0
            sheet = wb.active

            z = 'a'
            seed = []
            # r=0;
            sfera = []
            weight = []
            weight2 = []
            weight3 = []
            weight4 = []
            weight5 = []
            weight6 = []
            sferaplebs = ['Бытовые', 'Военные', 'Исследовательские', 'Промышленные', 'Строительные', 'Сельскохозяйственные',
                      'Логистические']

            for x in map(chr, range(*map(ord, ['K', 'R']))):  # С 20 по 21 позволяет выводить любой диапозон экселя
                for i in range(1, 2):
                    p = sheet[str(x) + str(i)].value
                    sfera.append(p)
            for x in map(chr, range(*map(ord, ['K', 'R']))):  # С 20 по 21 позволяет выводить любой диапозон экселя
                for i in range(2, 3):
                    p = sheet[str(x) + str(i)].value
                    p1 = float(p)
                    weight.append(p1)
            # print(sfera)
            # print(weight)

            rce = random.choices([1, 2, 3, 4, 5, 6, 7],
                                weights=[weight[0], weight[1], weight[2], weight[3], weight[4], weight[5], weight[6]])
            if rce[0] == 1:
                z = 'K'
            elif rce[0] == 2:
                z = 'L'
            elif rce[0] == 3:
                z = 'M'
            elif rce[0] == 4:
                z = 'N'
            elif rce[0] == 5:
                z = 'O'
            elif rce[0] == 6:
                z = 'P'
            elif rce[0] == 7:
                z = 'Q'

            seed.append(rce[0])
            for y in range(4, 12):
                r = sheet[z + str(y)].value
                weight2.append(r)

            rce = random.choices([1, 2, 3, 4, 5, 6, 7, 8],
                                weights=[weight2[0], weight2[1], weight2[2], weight2[3], weight2[4], weight2[5], weight2[6],
                                      weight2[7]])

            seed.append(rce[0])
            for y1 in range(13, 20):
                r = sheet[z + str(y1)].value
                weight3.append(r)

            rce = random.choices([1, 2, 3, 4, 5, 6, 7],
                                weights=[weight3[0], weight3[1], weight3[2], weight3[3], weight3[4], weight3[5], weight3[6]])

            seed.append(rce[0])
            for y2 in range(21, 24):
                r = sheet[z + str(y2)].value
                weight4.append(r)

            rce = random.choices([1, 2, 3], weights=[weight4[0], weight4[1], weight4[2]])

            seed.append(rce[0])
            for y3 in range(25, 28):
                r = sheet[z + str(y3)].value
                weight5.append(r)

            rce = random.choices([1, 2, 3], weights=[weight5[0], weight5[1], weight5[2]])

            seed.append(rce[0])
            for y4 in range(29, 33):
                r = sheet[z + str(y4)].value
                weight6.append(r)

            rce = random.choices([1, 2, 3, 4], weights=[weight6[0], weight6[1], weight6[2], weight6[3]])

            seed.append(rce[0])
            #print(seed)
            # ниже переводим сгенерированны сид в int число
            rez = ''
            rezz = 0
            f = len(seed)
            for i in range(f):
                rez += str(seed[i])
                rezz = int(rez)
            sovpad.append( rezz ) # добавляем числовой сид в массив
            #print(rez) #
            worksheet.write( len(sovpad), 1, rezz ) # записываем сгенерированный сид в эксель
            if len(sovpad) != len(set(sovpad)): # проверяем сид на совпадения путем сравнения длин массива с сидами и отсортированного массива без повторений сидов
                del sovpad[-1]  # удаляем последний сгенерированынй повторяющийся сид
                #print("D'oh!!!!!")

#workbook.close()

print([p for p, t in Counter(sovpad).items() if t > 1]) # если есть совпадения, выводит какие именно
print(len(sovpad)) # количество записанных сидов
print(surname) #выводит список фио
print(sovpad) # выводит список сгенерированных сидов
for q in range (0, 29):
    wb = openpyxl.reader.excel.load_workbook(filename="Robot.xlsx", data_only=True)
    wb.active = 0
    sheet = wb.active
    f=str(sovpad[q])
    if f[0:1] == "1":
        mint = sheet["K1"].value
    elif f[0:1] == "2":
        mint = sheet["L1"].value
    elif f[0:1] == "3":
        mint = sheet["M1"].value
    elif f[0:1] == "4":
        mint = sheet["N1"].value
    elif f[0:1] == "5":
        mint = sheet["O1"].value
    elif f[0:1] == "6":
        mint = sheet["P1"].value
    elif f[0:1] == "7":
        mint = sheet["Q1"].value
    shnaps.append(mint)
    if f[1:2] == "1":
        mint = sheet["J4"].value
    elif f[1:2] == "2":
        mint = sheet["J5"].value
    elif f[1:2] == "3":
        mint = sheet["J6"].value
    elif f[1:2] == "4":
        mint = sheet["J7"].value
    elif f[1:2] == "5":
        mint = sheet["J8"].value
    elif f[1:2] == "6":
        mint = sheet["J9"].value
    elif f[1:2] == "7":
        mint = sheet["J10"].value
    elif f[1:2] == "8":
        mint = sheet["J11"].value
    shnaps.append(mint)
    if f[2:3] == "1":
        mint = sheet["J13"].value
    elif f[2:3] == "2":
        mint = sheet["J14"].value
    elif f[2:3] == "3":
        mint = sheet["J15"].value
    elif f[2:3] == "4":
        mint = sheet["J16"].value
    elif f[2:3] == "5":
        mint = sheet["J17"].value
    elif f[2:3] == "6":
        mint = sheet["J18"].value
    elif f[2:3] == "7":
        mint = sheet["J19"].value
    shnaps.append(mint)
    if f[3:4] == "1":
        mint = sheet["J21"].value
    elif f[3:4] == "2":
        mint = sheet["J22"].value
    elif f[3:4] == "3":
        mint = sheet["J23"].value
    shnaps.append(mint)
    if f[4:5] == "1":
        mint = sheet["J25"].value
    elif f[4:5] == "2":
        mint = sheet["J26"].value
    elif f[4:5] == "3":
        mint = sheet["J27"].value
    shnaps.append(mint)
    if f[5:6] == "1":
        mint = sheet["J29"].value
    elif f[5:6] == "2":
        mint = sheet["J30"].value
    elif f[5:6] == "3":
        mint = sheet["J31"].value
    elif f[5:6] == "4":
        mint = sheet["J32"].value
    shnaps.append(mint)
    #print(shnaps)

    for x in range(6):
        worksheet.write(1+q, 2+x, shnaps[x])
    shnaps.clear()
shapka= ["ФИО","Seed","Сфера деятельности","Среда","Размер","Функция","Управление","Привод"]
for R in range(0, 8):
    worksheet.write(0, R, shapka[R])

workbook.close()